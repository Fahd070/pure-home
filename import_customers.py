#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pure Home — Customer Import Script
Reads مواعيد الصيانة sheet from Excel and imports customers + next-maintenance
appointments via the live API at https://wfm-system.onrender.com

Usage:
  python import_customers.py
  python import_customers.py --dry-run          # preview only, no writes
  python import_customers.py --api https://...  # override API base URL
"""

import sys, os, re, json, time, argparse, traceback
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    import urllib.request, urllib.error, urllib.parse
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install openpyxl")
    sys.exit(1)

# ─── Configuration ──────────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Users\fahd1\Desktop\عوض-جداول متابعة العملاء.xlsx"
SHEET_NAME  = "مواعيد الصيانة"
API_BASE    = "https://wfm-system.onrender.com"
ADMIN_EMAIL = "admin@wfm.local"
ADMIN_PASS  = "admin123"
DATA_START_ROW = 4   # first row with actual customer data
YELLOW_RGB  = "FFFFFF00"   # openpyxl RGB for yellow fill

# Delay between write operations (seconds) — keeps well under 500 req/15min
WRITE_DELAY = 3.0

# ─── Colour → city mapping ──────────────────────────────────────────────────
KNOWN_CITIES_IN_HOOD = {
    "مكة":      "مكة المكرمة",
    "الطائف":   "الطائف",
    "الطايف":   "الطائف",
    "المدينة":  "المدينة المنورة",
    "الرياض":   "الرياض",
    "ابها":     "أبها",
    "أبها":     "أبها",
    "القصيم":   "القصيم",
    "الاحساء":  "الأحساء",
}

# ─── Phone normalisation ─────────────────────────────────────────────────────

def normalize_phone(raw):
    """Return (primary_phone_str, list_of_all_valid, raw_str)."""
    if not raw:
        return None, [], str(raw)
    s = str(raw).strip()
    if s.endswith('.0'):
        s = s[:-2]
    parts = re.split(r'[/,،\-–]+', s)
    valid = []
    for part in parts:
        digits = re.sub(r'\D', '', part)
        if not digits:
            continue
        if len(digits) == 9 and digits[0] == '5':
            digits = '0' + digits
        if re.match(r'^05\d{8}$', digits):
            if digits not in valid:
                valid.append(digits)
    primary = valid[0] if valid else None
    return primary, valid, str(raw)

# ─── Location parser ─────────────────────────────────────────────────────────

def parse_location(hood_value, is_yellow):
    """Return (city, district, full_location_str)."""
    hood = str(hood_value).strip() if hood_value else ''
    if not is_yellow:
        city     = "جدة"
        district = hood if hood else "—"
        loc_str  = f"جدة - {district}" if district != "—" else "جدة"
        return city, district, loc_str

    city = None
    for prefix, city_name in KNOWN_CITIES_IN_HOOD.items():
        if hood.startswith(prefix):
            city = city_name
            district = hood[len(prefix):].strip(' -–•,')
            if not district:
                district = city_name
            loc_str = f"{city_name} - {district}"
            return city, district, loc_str

    city     = hood if hood else "—"
    district = hood if hood else "—"
    loc_str  = hood if hood else "—"
    return city, district, loc_str

# ─── Date helpers ─────────────────────────────────────────────────────────────

def to_iso(val):
    """Convert openpyxl cell value (datetime or string) to ISO date string or None."""
    if val is None:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    m = re.search(r'(\d{4}-\d{2}-\d{2})', s)
    if m:
        return m.group(1)
    return None

def is_future(iso_date):
    """True if iso_date (YYYY-MM-DD) is today or in the future."""
    if not iso_date:
        return False
    from datetime import date
    try:
        d = date.fromisoformat(iso_date)
        return d >= date.today()
    except:
        return False

# ─── HTTP helpers ─────────────────────────────────────────────────────────────

def api_call(method, path, token=None, body=None, timeout=90):
    url = API_BASE + path
    data = json.dumps(body).encode('utf-8') if body else None
    headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}
    if token:
        headers['Authorization'] = f'Bearer {token}'
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status, json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        body_bytes = e.read()
        try:
            err_body = json.loads(body_bytes.decode('utf-8'))
        except:
            err_body = {'raw': body_bytes.decode('utf-8', errors='replace')}
        return e.code, err_body
    except (urllib.error.URLError, TimeoutError, OSError) as e:
        return 0, {'error': 'CONNECTION_ERROR', 'message': str(e)}


def api_call_retry(method, path, token=None, body=None, timeout=90, max_retries=4):
    """api_call with automatic 429 back-off and connection retry."""
    for attempt in range(max_retries):
        if attempt > 0:
            wait = 60 * attempt
            print(f"\n    [retry {attempt}/{max_retries-1}] waiting {wait}s …", flush=True)
            time.sleep(wait)
        code, resp = api_call(method, path, token, body, timeout)
        if code == 429:
            print(f"\n    [429 rate-limited] will retry after back-off …", flush=True)
            time.sleep(90)
            continue
        if code == 0:
            print(f"\n    [connection error] {resp.get('message')} — retrying …", flush=True)
            continue
        return code, resp
    return code, resp


def get_token():
    for attempt in range(4):
        if attempt > 0:
            wait = 30 * attempt
            print(f"  Retrying auth in {wait}s (attempt {attempt+1}/4)…", flush=True)
            time.sleep(wait)
        print(f"  Authenticating as {ADMIN_EMAIL} …", end='', flush=True)
        code, resp = api_call('POST', '/api/auth/login', body={
            'email': ADMIN_EMAIL,
            'password': ADMIN_PASS,
        }, timeout=120)
        if code == 200 and resp.get('success'):
            print(" OK")
            return resp['data']['token'], resp['data']['user']['id']
        if code == 429:
            print(f" RATE LIMITED — waiting 60s before retry …", flush=True)
            time.sleep(60)
            continue
        if code == 0:
            print(f" CONNECTION ERROR: {resp.get('message')} — retrying …")
            continue
        print(f" FAILED (HTTP {code}): {resp}")
        break
    return None, None

# ─── Bulk-fetch existing customers ────────────────────────────────────────────

def bulk_fetch_existing_customers(token):
    """
    Fetch ALL customers from the API using pagination.
    Returns (by_phone: dict[str, customer], by_name: dict[str, customer]).
    This replaces per-row GET calls, cutting API usage from ~800 GETs to ~5-10.
    """
    by_phone = {}
    by_name  = {}
    page = 1
    limit = 100
    total_fetched = 0

    print("  Fetching existing customers from API …", flush=True)
    while True:
        code, resp = api_call_retry(
            'GET', f'/api/customers?limit={limit}&page={page}', token, timeout=60
        )
        if code != 200:
            print(f"  WARNING: Failed to fetch page {page} (HTTP {code}) — stopping bulk fetch")
            break

        customers = resp.get('data') or []
        if not customers:
            break

        for c in customers:
            phone = (c.get('phone') or '').strip()
            name  = (c.get('name')  or '').strip()
            if phone:
                by_phone[phone] = c
            if name:
                by_name[name] = c

        total_fetched += len(customers)
        meta = resp.get('meta') or resp.get('pagination') or {}
        total_count = meta.get('total') or 0
        total_pages = max(1, -(-total_count // limit))  # ceil division
        print(f"    Page {page}/{total_pages}: {len(customers)} customers (total in DB: {total_count})", flush=True)

        if page >= total_pages or len(customers) < limit:
            break
        page += 1
        time.sleep(0.5)  # gentle pause between paginated GETs

    print(f"  Total existing customers loaded: {total_fetched} "
          f"({len(by_phone)} unique phones, {len(by_name)} unique names)")
    return by_phone, by_name

# ─── Main import ─────────────────────────────────────────────────────────────

def main(dry_run=False, api_override=None):
    global API_BASE
    if api_override:
        API_BASE = api_override.rstrip('/')

    print("=" * 60)
    print("Pure Home — Customer Import")
    print(f"Excel : {EXCEL_PATH}")
    print(f"Sheet : {SHEET_NAME}")
    print(f"API   : {API_BASE}")
    print(f"Mode  : {'DRY RUN (no writes)' if dry_run else 'LIVE'}")
    print("=" * 60)

    # ── Load Excel ──
    print("\nLoading Excel file …")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]
    print(f"  Rows: {ws.max_row - DATA_START_ROW + 1} data rows (starting row {DATA_START_ROW})")

    # ── Authenticate ──
    if not dry_run:
        print("\nConnecting to API …")
        print("  Waking Render service …", end='', flush=True)
        code, _ = api_call('GET', '/health')
        print(f" {code}")

        token, admin_user_id = get_token()
        if not token:
            print("\nCannot proceed without valid auth token.")
            print("Check ADMIN_EMAIL / ADMIN_PASS at the top of this script.")
            sys.exit(1)

        # ── Bulk fetch existing customers (replaces per-row GET calls) ──
        print()
        existing_by_phone, existing_by_name = bulk_fetch_existing_customers(token)
    else:
        token = None
        admin_user_id = "dry-run"
        existing_by_phone, existing_by_name = {}, {}

    # ── Import loop ──
    stats = {
        'scanned': 0, 'created': 0, 'updated': 0,
        'skipped_dup': 0, 'skipped_noname': 0, 'skipped_nophone': 0,
        'appt_created': 0, 'errors': []
    }
    skipped_details = []

    print("\nImporting customers …\n")

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        name_raw  = ws.cell(row_num, 1).value
        inst_raw  = ws.cell(row_num, 2).value
        lastm_raw = ws.cell(row_num, 3).value
        nextm_raw = ws.cell(row_num, 4).value
        hood_raw  = ws.cell(row_num, 7).value
        imp_raw   = ws.cell(row_num, 8).value
        notes_raw = ws.cell(row_num, 9).value
        phone_raw = ws.cell(row_num, 10).value

        if not name_raw and not phone_raw:
            continue

        stats['scanned'] += 1
        name = str(name_raw).strip() if name_raw else ''

        if not name:
            stats['skipped_noname'] += 1
            skipped_details.append({'row': row_num, 'reason': 'empty name', 'phone': str(phone_raw)})
            continue

        primary_phone, all_phones, phone_str_raw = normalize_phone(phone_raw)
        if not primary_phone:
            stats['skipped_nophone'] += 1
            skipped_details.append({'row': row_num, 'reason': f'invalid phone: {repr(phone_raw)}', 'name': name})
            continue

        inst_iso  = to_iso(inst_raw)
        lastm_iso = to_iso(lastm_raw)
        nextm_iso = to_iso(nextm_raw)

        notes_parts = []
        if imp_raw:
            notes_parts.append(f"مهم: {str(imp_raw).strip()}")
        if notes_raw:
            notes_parts.append(str(notes_raw).strip())
        if lastm_iso:
            notes_parts.append(f"آخر صيانة: {lastm_iso}")
        if len(all_phones) > 1:
            extras = ' / '.join(all_phones[1:])
            notes_parts.append(f"أرقام إضافية: {extras}")
        notes_combined = "\n".join(notes_parts)[:2000]

        hood_cell = ws.cell(row_num, 7)
        fg = hood_cell.fill.fgColor
        is_yellow = (fg and fg.type == 'rgb' and fg.rgb == YELLOW_RGB)
        city, district, loc_str = parse_location(hood_raw, is_yellow)

        print(f"  R{row_num:4d}  {name[:35]:<35}  {primary_phone}  {loc_str[:30]}", end='', flush=True)

        if dry_run:
            print("  [DRY RUN]")
            stats['created'] += 1
            continue

        # ── Check duplicate using in-memory lookup (no API call) ──
        existing = existing_by_phone.get(primary_phone) or existing_by_name.get(name)

        if existing:
            stats['skipped_dup'] += 1
            print(f"  → DUP (id={existing['id'][:8]})")

            # Update missing fields without overwriting
            updates = {}
            if not existing.get('notes') and notes_combined:
                updates['notes'] = notes_combined
            if not existing.get('installationDate') and inst_iso:
                updates['installationDate'] = inst_iso

            if updates:
                code, resp = api_call_retry('PUT', f"/api/customers/{existing['id']}", token, {
                    **updates,
                    'version': existing.get('version', 1),
                })
                if code == 200:
                    stats['updated'] += 1
                    print(f"         → Updated: {list(updates.keys())}")
                    # Refresh in-memory entry
                    existing_by_phone[primary_phone] = resp.get('data', existing)
                    existing_by_name[name] = resp.get('data', existing)
                else:
                    print(f"         → Update failed: {resp}")
                time.sleep(WRITE_DELAY)

            # Create appointment if future date not yet in system
            if nextm_iso and is_future(nextm_iso):
                code2, appts_resp = api_call_retry(
                    'GET', f"/api/appointments?customerId={existing['id']}&limit=50", token
                )
                existing_dates = set()
                for a in (appts_resp.get('data') or []):
                    d = to_iso(a.get('scheduledDate', ''))
                    if d:
                        existing_dates.add(d)
                if nextm_iso not in existing_dates:
                    code3, appt_resp = api_call_retry('POST', '/api/appointments', token, {
                        'customerId': existing['id'],
                        'type': 'MAINTENANCE',
                        'scheduledDate': nextm_iso + 'T08:00:00.000Z',
                        'notes': 'موعد صيانة مستورد من جدول المتابعة',
                    })
                    if code3 == 201:
                        stats['appt_created'] += 1
                        print(f"         → Appointment created for {nextm_iso}")
                    else:
                        print(f"         → Appointment failed: {appt_resp}")
                    time.sleep(WRITE_DELAY)
            continue

        # ── Create new customer ──
        payload = {
            'name': name,
            'phone': primary_phone,
            'maintenanceCycle': 'MONTHLY',
            'maintenanceFrequency': 3,
            'notes': notes_combined if notes_combined else None,
            'installationDate': inst_iso,
            'address': {
                'city': city,
                'district': district,
                'street': loc_str,
            },
        }
        payload = {k: v for k, v in payload.items() if v is not None}

        code, resp = api_call_retry('POST', '/api/customers', token, payload)
        time.sleep(WRITE_DELAY)

        if code == 201:
            customer_id = resp['data']['id']
            stats['created'] += 1
            print(f"  → CREATED ({customer_id[:8]})")

            # Add to in-memory lookup to protect duplicates within this run
            new_cust = resp['data']
            existing_by_phone[primary_phone] = new_cust
            existing_by_name[name] = new_cust

            if nextm_iso and is_future(nextm_iso):
                code2, appt_resp = api_call_retry('POST', '/api/appointments', token, {
                    'customerId': customer_id,
                    'type': 'MAINTENANCE',
                    'scheduledDate': nextm_iso + 'T08:00:00.000Z',
                    'notes': 'موعد صيانة مستورد من جدول المتابعة',
                })
                if code2 == 201:
                    stats['appt_created'] += 1
                else:
                    print(f"         → Appt failed: {appt_resp}")
                time.sleep(WRITE_DELAY)

        elif code == 409:
            stats['skipped_dup'] += 1
            print(f"  → CONFLICT/DUP")
        else:
            stats['errors'].append({'row': row_num, 'name': name, 'code': code, 'resp': str(resp)[:200]})
            print(f"  → ERROR {code}: {str(resp)[:100]}")

    # ── Report ──────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("IMPORT REPORT")
    print("=" * 60)
    print(f"  Total rows scanned     : {stats['scanned']}")
    print(f"  Customers created      : {stats['created']}")
    print(f"  Customers updated      : {stats['updated']}")
    print(f"  Duplicates skipped     : {stats['skipped_dup']}")
    print(f"  Skipped (no name)      : {stats['skipped_noname']}")
    print(f"  Skipped (invalid phone): {stats['skipped_nophone']}")
    print(f"  Appointments created   : {stats['appt_created']}")
    print(f"  Errors                 : {len(stats['errors'])}")

    if skipped_details:
        print(f"\nSkipped records (first 20):")
        for s in skipped_details[:20]:
            print(f"  R{s['row']}: {s['reason']} | {s.get('name', s.get('phone', ''))}")

    if stats['errors']:
        print(f"\nError records:")
        for e in stats['errors'][:20]:
            print(f"  R{e['row']}: {e['name']} → HTTP {e['code']}: {e['resp'][:120]}")

    print("=" * 60)
    print(f"Done. {'(DRY RUN — nothing was written)' if dry_run else 'Import complete.'}")

    report_path = os.path.join(os.path.dirname(__file__), 'import_report.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump({
            'stats': {**stats, 'errors': stats['errors']},
            'skipped': skipped_details,
        }, f, ensure_ascii=False, indent=2)
    print(f"Report saved → {report_path}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Pure Home Customer Import')
    parser.add_argument('--dry-run', action='store_true', help='Preview only, no writes')
    parser.add_argument('--api', default=None, help='Override API base URL')
    args = parser.parse_args()
    main(dry_run=args.dry_run, api_override=args.api)
